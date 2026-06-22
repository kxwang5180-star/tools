from __future__ import annotations

from io import BytesIO, StringIO
from decimal import Decimal
from pathlib import Path
import re
import shutil
from typing import Any, Dict, Iterable, List

from .models import BillingBatch, CompanyQuarterBill, StoreFeeLine
from .parser import cents_to_amount


def build_template_rows(batch: BillingBatch, config: Dict[str, Any]) -> List[Dict[str, str]]:
    total_by_bill = {
        (bill.secondary_company, bill.quarter): cents_to_amount(bill.total_amount_cents)
        for bill in batch.bills
    }
    rows: List[Dict[str, str]] = []
    for line in batch.lines:
        rows.append(
            {
                "母公司": line.secondary_company,
                "二级公司": line.secondary_company,
                "季度": line.quarter,
                "合同编号": line.contract_number,
                "门店代码": line.store_code,
                "门店名称": line.store_name,
                "发票号码": line.invoice_number,
                "服务费金额": cents_to_amount(line.service_fee_cents),
                "单据合计": total_by_bill.get((line.secondary_company, line.quarter), "0.00"),
            }
        )
    return rows


def export_upload_template_xlsx(batch: BillingBatch, config: Dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError("导出 .xlsx 需要安装 openpyxl。") from exc

    if batch.template_file and _looks_like_finance_import_template(Path(batch.template_file)):
        return _export_finance_import_template_xlsx(batch, config, Path(batch.template_file))

    columns = config.get("upload_template_columns") or [
        "二级公司",
        "季度",
        "合同编号",
        "门店代码",
        "门店名称",
        "发票号码",
        "服务费金额",
        "单据合计",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "系统上传模板"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_index, value=column)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows = build_template_rows(batch, config)
    for row_index, row in enumerate(rows, start=2):
        for col_index, column in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=col_index, value=row.get(column, ""))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for col_index, column in enumerate(columns, start=1):
        width = min(max(len(column) + 4, 12), 24)
        for row in rows[:50]:
            width = min(max(width, len(str(row.get(column, ""))) + 2), 32)
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_upload_template_csv(batch: BillingBatch, config: Dict[str, Any]) -> str:
    import csv

    columns = config.get("upload_template_columns") or []
    rows = build_template_rows(batch, config)
    handle = StringIO()
    writer = csv.DictWriter(handle, fieldnames=columns)
    writer.writeheader()
    writer.writerows(rows)
    return handle.getvalue()


def export_company_detail_workbooks(
    bills: List[CompanyQuarterBill],
    lines: List[StoreFeeLine],
    target_dir: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if target_dir.exists():
        shutil.rmtree(target_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    lines_by_company: Dict[str, List[StoreFeeLine]] = {}
    for line in lines:
        lines_by_company.setdefault(line.secondary_company, []).append(line)

    for bill in bills:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "汇总"
        summary.append(["二级公司", bill.secondary_company])
        summary.append(["季度", bill.quarter])
        summary.append(["明细行数", bill.line_count])
        summary.append(["门店数量", len(bill.store_codes)])
        summary.append(["发票数量", len(bill.invoice_numbers)])
        summary.append(["金额合计", cents_to_amount(bill.total_amount_cents)])

        detail = workbook.create_sheet("门店明细")
        columns = ["门店代码", "门店名称", "合同编号", "发票号码", "服务费金额", "季度"]
        detail.append(columns)
        for line in sorted(lines_by_company.get(bill.secondary_company, []), key=lambda item: (item.store_code, item.store_name)):
            detail.append(
                [
                    line.store_code,
                    line.store_name,
                    line.contract_number,
                    line.invoice_number,
                    cents_to_amount(line.service_fee_cents),
                    line.quarter,
                ]
            )

        _style_company_workbook(workbook)
        workbook.save(target_dir / f"{_safe_file_name(bill.secondary_company)}_{_safe_file_name(bill.quarter)}.xlsx")


def _style_company_workbook(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2F3")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_index in range(1, sheet.max_column + 1):
            width = 10
            for row_index in range(1, min(sheet.max_row, 120) + 1):
                width = max(width, len(str(sheet.cell(row_index, column_index).value or "")) + 2)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(width, 42)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")


def _safe_file_name(value: str) -> str:
    return re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", value).strip("_") or "未命名"


def _looks_like_finance_import_template(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return False
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "数据" not in workbook.sheetnames:
        return False
    sheet = workbook["数据"]
    header = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    return {"leCode", "ccCode", "operationSubTypeCode", "standardCurrencyAmount"}.issubset(set(header))


def _export_finance_import_template_xlsx(
    batch: BillingBatch, config: Dict[str, Any], template_path: Path
) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(template_path)
    sheet = workbook["数据"]
    headers = [cell.value for cell in sheet[1]]
    header_index = {
        str(value).strip(): index + 1 for index, value in enumerate(headers) if value is not None
    }
    if sheet.max_row > 2:
        sheet.delete_rows(3, sheet.max_row - 2)

    defaults = config.get("finance_import_defaults", {})
    for row_index, row in enumerate(_finance_import_rows(batch, defaults), start=3):
        for key, value in row.items():
            column = header_index.get(key)
            if column:
                sheet.cell(row=row_index, column=column, value=value)

    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = max(
            sheet.column_dimensions[get_column_letter(column)].width or 10,
            14,
        )
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _finance_import_rows(batch: BillingBatch, defaults: Dict[str, Any]) -> List[Dict[str, Any]]:
    suffix = str(defaults.get("ccCode_suffix", "200002"))
    rows: List[Dict[str, Any]] = []
    for bill in sorted(batch.bills, key=lambda item: (_bill_company_code(item), item.secondary_company)):
        le_code = _bill_company_code(bill)
        rows.append(
            {
                "leCode": le_code,
                "travelerEmpNo": None,
                "ccCode": f"{le_code}{suffix}",
                "operationSubTypeCode": defaults.get("operationSubTypeCode", "hdl0085-001"),
                "standardCurrencyAmount": float((Decimal(bill.total_amount_cents) / Decimal(100)).quantize(Decimal("0.01"))),
                "taxType": defaults.get("taxType", "001105"),
                "assetsMainNo": None,
                "assetsChildNo": None,
                "assetsCount": None,
                "vendorCode": defaults.get("vendorCode"),
            }
        )
    return rows


def _bill_company_code(bill: Any) -> str:
    match = re.match(r"^\s*(\d+)\b", bill.secondary_company)
    if match:
        return match.group(1)
    return bill.store_codes[0] if bill.store_codes else bill.secondary_company.split(" ", 1)[0]
