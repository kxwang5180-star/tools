from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional


@dataclass(frozen=True)
class CompanyStructureMatch:
    company_code: str
    short_name: str
    secondary_code: str
    secondary_name: str
    category: str
    source_sheet: str

    @property
    def company_label(self) -> str:
        return f"{self.secondary_code} {self.secondary_name}".strip()


@dataclass
class CompanyStructureLookup:
    by_code: Dict[str, CompanyStructureMatch]
    by_name: Dict[str, List[CompanyStructureMatch]]
    source_description: str

    def find(self, store_code: str, store_name: str) -> Optional[CompanyStructureMatch]:
        if store_code and store_code in self.by_code:
            return self.by_code[store_code]
        matches = self.by_name.get(store_name, []) if store_name else []
        if len(matches) == 1:
            return matches[0]
        preferred_categories = ["火锅门店", "每客美餐", "独立品牌", "创收项目", "优鼎优门店"]
        for category in preferred_categories:
            preferred = [match for match in matches if match.category == category]
            if len(preferred) == 1:
                return preferred[0]
        return None


def load_company_structure_lookup(path: Path) -> CompanyStructureLookup:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("读取公司架构表需要安装 openpyxl。") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    if "账单主体匹配" in workbook.sheetnames:
        return _load_from_matching_workbook(workbook, path)
    if "公司代码（国内+海外）" in workbook.sheetnames:
        return _load_from_structure_workbook(workbook, path)
    raise ValueError("公司架构/匹配表缺少可识别的 sheet：账单主体匹配 或 公司代码（国内+海外）。")


def remap_lines_to_secondary_company(
    lines: List[Any], structure_file: Path
) -> tuple[List[Any], List[Any], str]:
    from .models import ValidationIssue

    lookup = load_company_structure_lookup(structure_file)
    issues: List[ValidationIssue] = []
    for line in lines:
        match = lookup.find(line.store_code, line.store_name)
        if match is None:
            issues.append(
                ValidationIssue(
                    severity="error",
                    code="company_structure_mapping_missing",
                    message=f"门店未匹配到二级公司：{line.store_name}（公司代码：{line.store_code or '空'}）。",
                    row_number=line.source_row,
                    company=line.secondary_company,
                )
            )
            continue
        line.secondary_company = match.company_label
    return lines, issues, lookup.source_description


def _load_from_matching_workbook(workbook: Any, path: Path) -> CompanyStructureLookup:
    rows: List[CompanyStructureMatch] = []
    for sheet_name in ["账单主体匹配", "门店项目匹配"]:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header = [_clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: i for i, name in enumerate(header) if name}
        required = {"公司代码", "门店/主体简称", "二级公司代码"}
        if not required.issubset(index):
            continue
        for raw in sheet.iter_rows(min_row=2, values_only=True):
            company_code = _clean(raw[index["公司代码"]])
            secondary_code = _clean(raw[index["二级公司代码"]])
            if not company_code or not secondary_code:
                continue
            rows.append(
                CompanyStructureMatch(
                    company_code=company_code,
                    short_name=_cell(raw, index, "门店/主体简称"),
                    secondary_code=secondary_code,
                    secondary_name=_cell(raw, index, "二级公司名称"),
                    category=_cell(raw, index, "分类大类"),
                    source_sheet=sheet_name,
                )
            )
    return _build_lookup(rows, f"{path.name}:账单主体匹配")


def _load_from_structure_workbook(workbook: Any, path: Path) -> CompanyStructureLookup:
    sheet = workbook["公司代码（国内+海外）"]
    header = [_clean(cell) for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    index = {name: i for i, name in enumerate(header) if name}
    required = {"公司代码", "简称", "二级公司"}
    if not required.issubset(index):
        raise ValueError("公司代码（国内+海外）缺少公司代码、简称或二级公司列。")

    source_rows: List[dict[str, str]] = []
    code_to_name: Dict[str, str] = {}
    for raw in sheet.iter_rows(min_row=3, values_only=True):
        company_code = _clean(raw[index["公司代码"]])
        if not company_code:
            continue
        row = {
            "company_code": company_code,
            "short_name": _clean(raw[index["简称"]]),
            "company_name": _cell(raw, index, "公司名称"),
            "license_name": _cell(raw, index, "营业执照"),
            "secondary_code": _clean(raw[index["二级公司"]]),
            "category": _cell(raw, index, "分类-大类"),
        }
        source_rows.append(row)
        code_to_name[company_code] = row["company_name"] or row["license_name"] or row["short_name"]

    rows = [
        CompanyStructureMatch(
            company_code=row["company_code"],
            short_name=row["short_name"],
            secondary_code=row["secondary_code"],
            secondary_name=code_to_name.get(row["secondary_code"], ""),
            category=row["category"],
            source_sheet="公司代码（国内+海外）",
        )
        for row in source_rows
        if row["secondary_code"]
    ]
    rows.extend(_load_youdingyou_rows(workbook, code_to_name))
    return _build_lookup(rows, f"{path.name}:公司代码（国内+海外）")


def _load_youdingyou_rows(workbook: Any, code_to_name: Dict[str, str]) -> List[CompanyStructureMatch]:
    rows: List[CompanyStructureMatch] = []
    secondary_name = code_to_name.get("8150", "北京优鼎优餐饮管理有限公司")
    if "门店基础表-优鼎优" in workbook.sheetnames:
        sheet = workbook["门店基础表-优鼎优"]
        header = [_clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: i for i, name in enumerate(header) if name}
        for raw in sheet.iter_rows(min_row=2, values_only=True):
            company_code = _cell(raw, index, "公司代码")
            short_name = _cell(raw, index, "店名") or _cell(raw, index, "门店名称")
            if not company_code or not short_name:
                continue
            rows.append(
                CompanyStructureMatch(
                    company_code=company_code,
                    short_name=short_name,
                    secondary_code="8150",
                    secondary_name=secondary_name,
                    category="优鼎优门店",
                    source_sheet="门店基础表-优鼎优",
                )
            )
    if "优鼎优（金蝶）" in workbook.sheetnames:
        sheet = workbook["优鼎优（金蝶）"]
        header = [_clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: i for i, name in enumerate(header) if name}
        for raw in sheet.iter_rows(min_row=2, values_only=True):
            company_code = _cell(raw, index, "金蝶代码")
            short_name = _cell(raw, index, "店名")
            if not company_code or not short_name:
                continue
            rows.append(
                CompanyStructureMatch(
                    company_code=company_code,
                    short_name=short_name,
                    secondary_code="8150",
                    secondary_name=secondary_name,
                    category="优鼎优门店",
                    source_sheet="优鼎优（金蝶）",
                )
            )
    return rows


def _build_lookup(rows: List[CompanyStructureMatch], source_description: str) -> CompanyStructureLookup:
    by_code: Dict[str, CompanyStructureMatch] = {}
    by_name: Dict[str, List[CompanyStructureMatch]] = defaultdict(list)
    for row in rows:
        by_code.setdefault(row.company_code, row)
        if row.short_name:
            if row not in by_name[row.short_name]:
                by_name[row.short_name].append(row)
    return CompanyStructureLookup(by_code=by_code, by_name=dict(by_name), source_description=source_description)


def _clean(value: Any) -> str:
    if value is None or value == -1:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell(row: tuple[Any, ...], index: Dict[str, int], key: str) -> str:
    position = index.get(key)
    if position is None or position >= len(row):
        return ""
    return _clean(row[position])
