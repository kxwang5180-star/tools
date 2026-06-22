from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from .models import StoreFeeLine, ValidationIssue


@dataclass
class ParsedRows:
    lines: List[StoreFeeLine]
    issues: List[ValidationIssue]
    detected_headers: Dict[str, str]


@dataclass
class StoreMasterRecord:
    store_code: str
    billing_name: str


def parse_service_fee_file(
    file_path: Path, config: Dict[str, Any], default_quarter: str
) -> ParsedRows:
    if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
        special = _try_parse_haidilao_payment_workbook(file_path, default_quarter)
        if special is not None:
            return special

    rows = _load_rows(file_path)
    if not rows:
        return ParsedRows(
            lines=[],
            issues=[
                ValidationIssue(
                    severity="error",
                    code="empty_source_file",
                    message="服务费文件没有可读取的数据。",
                )
            ],
            detected_headers={},
        )

    header_index, headers = _find_header_row(rows, config.get("field_aliases", {}))
    if header_index is None or not headers:
        return ParsedRows(
            lines=[],
            issues=[
                ValidationIssue(
                    severity="error",
                    code="header_not_found",
                    message="未找到固定模板表头，请检查服务费 Excel 字段配置。",
                )
            ],
            detected_headers={},
        )

    field_to_header, header_issues = _map_headers(headers, config)
    issues = list(header_issues)
    lines: List[StoreFeeLine] = []

    for offset, raw_row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row = _row_to_dict(headers, raw_row)
        if _is_empty_row(row.values()):
            continue
        line, row_issues = _parse_line(row, field_to_header, offset, default_quarter)
        issues.extend(row_issues)
        if line is not None:
            lines.append(line)

    issues.extend(_find_duplicate_store_warnings(lines, config))
    return ParsedRows(lines=lines, issues=issues, detected_headers=field_to_header)


def cents_to_amount(cents: int) -> str:
    return f"{Decimal(cents) / Decimal(100):.2f}"


def parse_amount_to_cents(value: Any) -> Tuple[Optional[int], Optional[str]]:
    if value is None or str(value).strip() == "":
        return None, "金额为空。"
    text = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    text = re.sub(r"\s+", "", text)
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None, f"金额格式无法识别：{value}"
    cents = int((amount * Decimal(100)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if cents < 0:
        return None, f"金额不能为负数：{value}"
    return cents, None


def _load_rows(file_path: Path) -> List[List[Any]]:
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_rows(file_path)
    if suffix in {".csv", ".txt"}:
        return _load_delimited_rows(file_path, ",")
    if suffix == ".tsv":
        return _load_delimited_rows(file_path, "\t")
    raise ValueError(f"不支持的服务费文件类型：{suffix}")


def _load_xlsx_rows(file_path: Path) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("读取 .xlsx 需要安装 openpyxl。") from exc

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: List[List[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_clean_cell(cell) for cell in row])
    return rows


def _try_parse_haidilao_payment_workbook(
    file_path: Path, default_quarter: str
) -> Optional[ParsedRows]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("读取 .xlsx 需要安装 openpyxl。") from exc

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    payment_sheets = [
        sheet for sheet in workbook.worksheets if sheet.title.replace(" ", "").startswith("付款明细")
    ]
    if not payment_sheets or "火锅门店" not in workbook.sheetnames:
        return None

    store_master_by_name = _load_store_master_map(workbook["火锅门店"])
    parent_company_code_by_store = (
        _load_parent_company_code_map(workbook["所属公司"])
        if "所属公司" in workbook.sheetnames
        else {}
    )
    parent_company_name_by_code = _infer_parent_company_names(
        store_master_by_name, parent_company_code_by_store
    )
    parent_company_code_by_billing_name = _infer_parent_company_codes_by_billing_name(
        store_master_by_name, parent_company_code_by_store
    )
    issues: List[ValidationIssue] = []
    lines: List[StoreFeeLine] = []
    detected_headers = {
        "secondary_company": "所属公司!母公司 + 火锅门店!开票名称",
        "quarter": default_quarter,
        "store_code": "火锅门店!公司代码",
        "store_name": "门店",
        "invoice_number": "发票号码",
        "service_fee": "2026年1-3月服务费",
    }

    for sheet in payment_sheets:
        columns = _detect_haidilao_payment_columns(sheet)
        missing = [key for key in ["store_name", "service_fee", "invoice_number", "billing_name"] if key not in columns]
        if missing:
            issues.append(
                ValidationIssue(
                    severity="error",
                    code="payment_sheet_columns_missing",
                    message=f"{sheet.title} 缺少必要列：{', '.join(missing)}",
                )
            )
            continue

        for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            store_name = _cell_text(_row_value(row, columns["store_name"]))
            amount_value = _row_value(row, columns["service_fee"])
            invoice_number = _cell_text(_row_value(row, columns["invoice_number"]))
            billing_name = _cell_text(_row_value(row, columns["billing_name"]))

            if not store_name and amount_value in (None, "") and not invoice_number:
                continue
            store_record = store_master_by_name.get(store_name) if store_name else None
            store_code = store_record.store_code if store_record else ""
            if amount_value in (None, "") and not invoice_number:
                continue
            if not invoice_number and not store_code:
                continue

            amount_cents, amount_error = parse_amount_to_cents(amount_value)
            if amount_error:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        code="invalid_service_fee",
                        message=f"{sheet.title} 第 {row_number} 行：{amount_error}",
                        row_number=row_number,
                    )
                )
                continue

            if not store_name:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        code="missing_store_name",
                        message=f"{sheet.title} 第 {row_number} 行缺少门店。",
                        row_number=row_number,
                    )
                )
                continue
            if not store_code:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        code="missing_store_code_mapping",
                        message=f"{sheet.title} 第 {row_number} 行门店没有匹配到公司代码：{store_name}",
                        row_number=row_number,
                    )
                )
                continue
            if not invoice_number:
                issues.append(
                    ValidationIssue(
                        severity="error",
                        code="missing_invoice_number",
                        message=f"{sheet.title} 第 {row_number} 行缺少发票号码。",
                        row_number=row_number,
                        company=billing_name,
                    )
                )
                continue
            if not billing_name:
                billing_name = store_record.billing_name if store_record else store_name
            secondary_company = _haidilao_secondary_company(
                store_name=store_name,
                store_code=store_code,
                billing_name=billing_name,
                parent_company_code_by_store=parent_company_code_by_store,
                parent_company_code_by_billing_name=parent_company_code_by_billing_name,
                parent_company_name_by_code=parent_company_name_by_code,
            )

            lines.append(
                StoreFeeLine(
                    source_row=row_number,
                    secondary_company=secondary_company,
                    quarter=default_quarter,
                    store_code=store_code,
                    store_name=store_name,
                    contract_number="",
                    invoice_number=invoice_number,
                    service_fee_cents=amount_cents or 0,
                )
            )

    issues.extend(_find_duplicate_store_warnings(lines, {"validation": {"warn_on_duplicate_store_in_company_quarter": True}}))
    return ParsedRows(lines=lines, issues=issues, detected_headers=detected_headers)


def _load_store_master_map(sheet: Any) -> Dict[str, StoreMasterRecord]:
    header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    header_index = {str(value).strip(): index for index, value in enumerate(header) if value}
    name_index = header_index.get("店名")
    code_index = header_index.get("公司代码")
    billing_index = header_index.get("开票名称")
    if name_index is None or code_index is None:
        return {}
    mapping: Dict[str, StoreMasterRecord] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = _cell_text(_row_value(row, name_index))
        code = _cell_text(_row_value(row, code_index))
        billing_name = _cell_text(_row_value(row, billing_index)) if billing_index is not None else ""
        if name and code:
            mapping[name] = StoreMasterRecord(store_code=code, billing_name=billing_name)
    return mapping


def _load_parent_company_code_map(sheet: Any) -> Dict[str, str]:
    header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    header_index = {str(value).strip(): index for index, value in enumerate(header) if value}
    name_index = header_index.get("简称")
    code_index = header_index.get("母公司")
    if name_index is None or code_index is None:
        return {}
    mapping: Dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = _cell_text(_row_value(row, name_index))
        code = _cell_text(_row_value(row, code_index))
        if name and code:
            mapping[name] = code
    return mapping


def _infer_parent_company_names(
    store_master_by_name: Dict[str, StoreMasterRecord],
    parent_company_code_by_store: Dict[str, str],
) -> Dict[str, str]:
    names_by_code: Dict[str, Dict[str, int]] = {}
    for store_name, parent_code in parent_company_code_by_store.items():
        record = store_master_by_name.get(store_name)
        if not record or not record.billing_name:
            continue
        names_by_code.setdefault(parent_code, {})
        names_by_code[parent_code][record.billing_name] = (
            names_by_code[parent_code].get(record.billing_name, 0) + 1
        )
    return {
        parent_code: max(name_counts.items(), key=lambda item: (item[1], len(item[0])))[0]
        for parent_code, name_counts in names_by_code.items()
        if name_counts
    }


def _infer_parent_company_codes_by_billing_name(
    store_master_by_name: Dict[str, StoreMasterRecord],
    parent_company_code_by_store: Dict[str, str],
) -> Dict[str, str]:
    codes_by_billing_name: Dict[str, Dict[str, int]] = {}
    for store_name, record in store_master_by_name.items():
        parent_code = parent_company_code_by_store.get(store_name)
        if not parent_code or not record.billing_name:
            continue
        codes_by_billing_name.setdefault(record.billing_name, {})
        codes_by_billing_name[record.billing_name][parent_code] = (
            codes_by_billing_name[record.billing_name].get(parent_code, 0) + 1
        )
    return {
        billing_name: max(code_counts.items(), key=lambda item: item[1])[0]
        for billing_name, code_counts in codes_by_billing_name.items()
        if code_counts
    }


def _haidilao_secondary_company(
    store_name: str,
    store_code: str,
    billing_name: str,
    parent_company_code_by_store: Dict[str, str],
    parent_company_code_by_billing_name: Dict[str, str],
    parent_company_name_by_code: Dict[str, str],
) -> str:
    parent_code = parent_company_code_by_store.get(store_name) or parent_company_code_by_billing_name.get(billing_name)
    if parent_code:
        parent_name = parent_company_name_by_code.get(parent_code) or billing_name or store_name
        return f"{parent_code} {parent_name}"
    if store_code and billing_name:
        return f"{store_code} {billing_name}"
    return billing_name or store_name


def _detect_haidilao_payment_columns(sheet: Any) -> Dict[str, int]:
    first_row = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    second_row = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    columns: Dict[str, int] = {}
    for index, value in enumerate(first_row):
        text = _cell_text(value)
        if text == "门店":
            columns["store_name"] = index
        elif text == "2026年1-3月服务费":
            columns["service_fee"] = index
        elif text == "开票名称":
            columns["billing_name"] = index

    invoice_candidates: List[int] = []
    for index, value in enumerate(first_row):
        if _cell_text(value) == "发票号码":
            invoice_candidates.append(index)
    for index, value in enumerate(second_row):
        if _cell_text(value) == "发票号码":
            invoice_candidates.append(index)
    if invoice_candidates:
        columns["invoice_number"] = max(invoice_candidates, key=lambda idx: _invoice_like_count(sheet, idx))
    return columns


def _invoice_like_count(sheet: Any, column_index: int) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=3, max_row=min(sheet.max_row, 80), values_only=True):
        value = _cell_text(_row_value(row, column_index))
        if re.fullmatch(r"\d{8,20}", value):
            count += 1
    return count


def _row_value(row: List[Any] | Tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_delimited_rows(file_path: Path, delimiter: str) -> List[List[Any]]:
    rows: List[List[Any]] = []
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row in reader:
            rows.append([_clean_cell(cell) for cell in row])
    return rows


def _find_header_row(
    rows: List[List[Any]], field_aliases: Dict[str, List[str]]
) -> Tuple[Optional[int], List[str]]:
    alias_set = {
        _normalize_header(alias)
        for aliases in field_aliases.values()
        for alias in aliases
    }
    best_index: Optional[int] = None
    best_headers: List[str] = []
    best_score = 0
    for index, row in enumerate(rows[:20]):
        headers = [str(cell).strip() if cell is not None else "" for cell in row]
        score = sum(1 for header in headers if _normalize_header(header) in alias_set)
        if score > best_score:
            best_index = index
            best_headers = headers
            best_score = score
    if best_score < 3:
        return None, []
    return best_index, best_headers


def _map_headers(headers: List[str], config: Dict[str, Any]) -> Tuple[Dict[str, str], List[ValidationIssue]]:
    normalized_to_actual = {
        _normalize_header(header): header.strip()
        for header in headers
        if header and str(header).strip()
    }
    mapped: Dict[str, str] = {}
    issues: List[ValidationIssue] = []
    for field, aliases in config.get("field_aliases", {}).items():
        for alias in aliases:
            actual = normalized_to_actual.get(_normalize_header(alias))
            if actual:
                mapped[field] = actual
                break

    for field in config.get("required_fields", []):
        if field not in mapped:
            issues.append(
                ValidationIssue(
                    severity="error",
                    code="missing_required_column",
                    message=f"缺少必填字段列：{field}",
                )
            )
    return mapped, issues


def _parse_line(
    row: Dict[str, Any],
    field_to_header: Dict[str, str],
    source_row: int,
    default_quarter: str,
) -> Tuple[Optional[StoreFeeLine], List[ValidationIssue]]:
    issues: List[ValidationIssue] = []

    def value(field: str) -> str:
        header = field_to_header.get(field)
        if not header:
            return ""
        raw = row.get(header, "")
        return str(raw).strip() if raw is not None else ""

    amount_cents, amount_error = parse_amount_to_cents(value("service_fee"))
    if amount_error:
        issues.append(
            ValidationIssue(
                severity="error",
                code="invalid_service_fee",
                message=amount_error,
                row_number=source_row,
            )
        )

    required_values = {
        "secondary_company": value("secondary_company"),
        "store_code": value("store_code"),
        "contract_number": value("contract_number"),
        "invoice_number": value("invoice_number"),
    }
    for field, field_value in required_values.items():
        if not field_value:
            issues.append(
                ValidationIssue(
                    severity="error",
                    code="missing_required_value",
                    message=f"第 {source_row} 行缺少必填值：{field}",
                    row_number=source_row,
                )
            )

    if amount_cents is None or any(not item for item in required_values.values()):
        return None, issues

    return (
        StoreFeeLine(
            source_row=source_row,
            secondary_company=value("secondary_company"),
            quarter=value("quarter") or default_quarter,
            store_code=value("store_code"),
            store_name=value("store_name"),
            contract_number=value("contract_number"),
            invoice_number=value("invoice_number"),
            service_fee_cents=amount_cents,
        ),
        issues,
    )


def _find_duplicate_store_warnings(
    lines: List[StoreFeeLine], config: Dict[str, Any]
) -> List[ValidationIssue]:
    if not config.get("validation", {}).get("warn_on_duplicate_store_in_company_quarter", True):
        return []
    seen: Dict[Tuple[str, str, str], int] = {}
    issues: List[ValidationIssue] = []
    for line in lines:
        key = (line.secondary_company, line.quarter, line.store_code)
        if key in seen:
            issues.append(
                ValidationIssue(
                    severity="warning",
                    code="duplicate_store_in_company_quarter",
                    message=(
                        f"门店 {line.store_code} 在 {line.secondary_company} "
                        f"{line.quarter} 出现多行，已纳入汇总。"
                    ),
                    row_number=line.source_row,
                    company=line.secondary_company,
                )
            )
        seen[key] = line.source_row
    return issues


def _row_to_dict(headers: List[str], raw_row: List[Any]) -> Dict[str, Any]:
    output: Dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        output[header] = raw_row[index] if index < len(raw_row) else ""
    return output


def _is_empty_row(values: Iterable[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value).strip().lower())
