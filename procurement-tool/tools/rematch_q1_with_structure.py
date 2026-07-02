from __future__ import annotations

import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FEE_FILE = Path("/Users/kk/Downloads/26年一季度发票/2026年1-3月发票/2026年1-3月海底捞付款申请(含发票号).xlsx")
FINANCE_TEMPLATE = Path("/Users/kk/Downloads/统一结算报账单导入模板 -导入版.xlsx")
INVOICE_ROOT = Path("/Users/kk/Downloads/26年一季度发票/2026年1-3月发票")
MATCH_TABLE = ROOT / "outputs" / "法人架构账单主体二级公司匹配表.xlsx"
OUTPUT_REVIEW = ROOT / "outputs" / "2026Q1服务费_二级公司重新匹配结果.xlsx"
OUTPUT_IMPORT = ROOT / "outputs" / "2026Q1统一结算报账单导入模板_二级公司口径.xlsx"
OUTPUT_INVOICES = ROOT / "outputs" / "2026Q1发票_按二级公司整理"
QUARTER = "2026 Q1"
INVOICE_SUFFIXES = {".pdf", ".ofd", ".jpg", ".jpeg", ".png", ".tif", ".tiff"}


@dataclass
class StructureMatch:
    company_code: str
    short_name: str
    secondary_code: str
    secondary_name: str
    category: str
    match_source: str


@dataclass
class FeeLine:
    sheet_name: str
    row_number: int
    store_name: str
    store_code: str
    billing_name: str
    invoice_number: str
    amount_cents: int
    match: Optional[StructureMatch]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_amount_to_cents(value: Any) -> int:
    text = clean(value).replace(",", "").replace("￥", "").replace("¥", "")
    if not text:
        return 0
    amount = Decimal(text)
    return int((amount * Decimal(100)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def cents_to_amount(cents: int) -> str:
    return f"{Decimal(cents) / Decimal(100):.2f}"


def load_structure_matches(path: Path) -> tuple[Dict[str, StructureMatch], Dict[str, List[StructureMatch]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["账单主体匹配"]
    headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: i for i, header in enumerate(headers)}
    by_code: Dict[str, StructureMatch] = {}
    by_name: Dict[str, List[StructureMatch]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        company_code = clean(row[index["公司代码"]])
        short_name = clean(row[index["门店/主体简称"]])
        secondary_code = clean(row[index["二级公司代码"]])
        if not company_code or not secondary_code:
            continue
        match = StructureMatch(
            company_code=company_code,
            short_name=short_name,
            secondary_code=secondary_code,
            secondary_name=clean(row[index["二级公司名称"]]),
            category=clean(row[index["分类大类"]]),
            match_source=clean(row[index["来源sheet"]]),
        )
        by_code[company_code] = match
        if short_name:
            by_name[short_name].append(match)
    return by_code, by_name


def load_store_master(workbook: Any) -> Dict[str, Dict[str, str]]:
    sheet = workbook["火锅门店"]
    headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: i for i, header in enumerate(headers)}
    output: Dict[str, Dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = clean(row[index["店名"]])
        if not name:
            continue
        output[name] = {
            "store_code": clean(row[index["公司代码"]]),
            "billing_name": clean(row[index["开票名称"]]),
        }
    return output


def detect_payment_columns(sheet: Any) -> Dict[str, int]:
    first = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    second = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    columns: Dict[str, int] = {}
    for index, value in enumerate(first):
        text = clean(value)
        if text == "门店":
            columns["store_name"] = index
        elif text == "2026年1-3月服务费":
            columns["amount"] = index
        elif text == "开票名称":
            columns["billing_name"] = index
    candidates: List[int] = []
    for index, value in enumerate(first):
        if clean(value) == "发票号码":
            candidates.append(index)
    for index, value in enumerate(second):
        if clean(value) == "发票号码":
            candidates.append(index)
    if candidates:
        columns["invoice_number"] = max(candidates, key=lambda column: invoice_like_count(sheet, column))
    missing = {"store_name", "amount", "billing_name", "invoice_number"} - set(columns)
    if missing:
        raise ValueError(f"{sheet.title} 缺少列：{', '.join(sorted(missing))}")
    return columns


def invoice_like_count(sheet: Any, column_index: int) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=3, max_row=min(sheet.max_row, 80), values_only=True):
        if re.fullmatch(r"\d{8,20}", clean(row[column_index] if column_index < len(row) else "")):
            count += 1
    return count


def find_match(
    store_code: str,
    store_name: str,
    by_code: Dict[str, StructureMatch],
    by_name: Dict[str, List[StructureMatch]],
) -> Optional[StructureMatch]:
    if store_code in by_code:
        return by_code[store_code]
    name_matches = by_name.get(store_name, [])
    if len(name_matches) == 1:
        return name_matches[0]
    if name_matches:
        hotpot = [match for match in name_matches if match.category == "火锅门店"]
        if len(hotpot) == 1:
            return hotpot[0]
    return None


def parse_fee_lines(
    source_file: Path,
    by_code: Dict[str, StructureMatch],
    by_name: Dict[str, List[StructureMatch]],
) -> List[FeeLine]:
    workbook = load_workbook(source_file, data_only=True, read_only=True)
    store_master = load_store_master(workbook)
    lines: List[FeeLine] = []
    payment_sheets = [sheet for sheet in workbook.worksheets if sheet.title.replace(" ", "").startswith("付款明细")]
    for sheet in payment_sheets:
        columns = detect_payment_columns(sheet)
        for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            store_name = clean(row[columns["store_name"]] if columns["store_name"] < len(row) else "")
            invoice_number = clean(row[columns["invoice_number"]] if columns["invoice_number"] < len(row) else "")
            amount_raw = row[columns["amount"]] if columns["amount"] < len(row) else None
            if not store_name and not invoice_number and amount_raw in (None, ""):
                continue
            if amount_raw in (None, "") and not invoice_number:
                continue
            master = store_master.get(store_name, {})
            store_code = master.get("store_code", "")
            billing_name = clean(row[columns["billing_name"]] if columns["billing_name"] < len(row) else "") or master.get("billing_name", "")
            if not invoice_number and not store_code:
                continue
            lines.append(
                FeeLine(
                    sheet_name=sheet.title,
                    row_number=row_number,
                    store_name=store_name,
                    store_code=store_code,
                    billing_name=billing_name,
                    invoice_number=invoice_number,
                    amount_cents=parse_amount_to_cents(amount_raw),
                    match=find_match(store_code, store_name, by_code, by_name),
                )
            )
    return lines


def map_invoice_files(root: Path) -> Dict[str, List[Path]]:
    output: Dict[str, List[Path]] = defaultdict(list)
    for file_path in root.rglob("*"):
        if not file_path.is_file() or file_path.suffix.lower() not in INVOICE_SUFFIXES or file_path.name.startswith("._"):
            continue
        match = re.search(r"\d{8,20}", file_path.name)
        if match:
            output[match.group(0)].append(file_path)
    return output


def safe_name(value: str) -> str:
    return re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", value).strip("_") or "未匹配"


def unique_path(target: Path) -> Path:
    if not target.exists():
        return target
    counter = 2
    while True:
        candidate = target.with_name(f"{target.stem}_{counter}{target.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def organize_invoices(lines: List[FeeLine], invoice_files: Dict[str, List[Path]], output_dir: Path) -> int:
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    copied: set[Path] = set()
    for line in lines:
        if not line.match or not line.invoice_number:
            continue
        folder = output_dir / safe_name(f"{line.match.secondary_code}_{line.match.secondary_name}_{QUARTER}")
        folder.mkdir(parents=True, exist_ok=True)
        for source in invoice_files.get(line.invoice_number, []):
            if source in copied:
                continue
            shutil.copy2(source, unique_path(folder / source.name))
            copied.add(source)
    return len(copied)


def aggregate(lines: Iterable[FeeLine]) -> List[Dict[str, Any]]:
    groups: Dict[tuple[str, str], List[FeeLine]] = defaultdict(list)
    for line in lines:
        if not line.match:
            groups[("未匹配", "未匹配")].append(line)
        else:
            groups[(line.match.secondary_code, line.match.secondary_name)].append(line)
    rows: List[Dict[str, Any]] = []
    for (code, name), group in sorted(groups.items(), key=lambda item: item[0]):
        rows.append(
            {
                "二级公司代码": code,
                "二级公司名称": name,
                "明细行数": len(group),
                "金额合计": cents_to_amount(sum(item.amount_cents for item in group)),
                "发票数量": len({item.invoice_number for item in group if item.invoice_number}),
                "门店数量": len({item.store_name for item in group if item.store_name}),
            }
        )
    return rows


def write_rows(sheet: Any, columns: List[str], rows: Iterable[Dict[str, Any]]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def style_workbook(workbook: Any) -> None:
    fill = PatternFill("solid", fgColor="DDEBF7")
    font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2F3")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, sheet.max_column + 1):
            width = 10
            for row in range(1, min(sheet.max_row, 120) + 1):
                width = max(width, len(str(sheet.cell(row, col).value or "")) + 2)
            sheet.column_dimensions[get_column_letter(col)].width = min(width, 42)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")


def export_review_workbook(lines: List[FeeLine], invoice_files: Dict[str, List[Path]], output_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "说明"
    matched_lines = [line for line in lines if line.match]
    missing_mapping = [line for line in lines if not line.match]
    missing_invoice = [line for line in lines if line.invoice_number and not invoice_files.get(line.invoice_number)]
    summary_rows = [
        ["文件", "2026Q1服务费_二级公司重新匹配结果"],
        ["费用来源", str(SOURCE_FEE_FILE)],
        ["匹配表", str(MATCH_TABLE)],
        ["匹配口径", "优先按门店公司代码匹配；公司代码缺失时按门店名称/主体简称匹配"],
        ["明细行数", len(lines)],
        ["已匹配行数", len(matched_lines)],
        ["未匹配行数", len(missing_mapping)],
        ["缺失发票文件行数", len(missing_invoice)],
        ["金额合计", cents_to_amount(sum(line.amount_cents for line in lines))],
    ]
    for row in summary_rows:
        summary.append(row)

    detail = workbook.create_sheet("门店明细")
    detail_columns = [
        "来源sheet",
        "来源行",
        "门店名称",
        "门店公司代码",
        "开票名称",
        "二级公司代码",
        "二级公司名称",
        "匹配来源",
        "分类",
        "发票号码",
        "发票文件数",
        "服务费金额",
        "匹配状态",
    ]
    detail_rows = []
    for line in lines:
        detail_rows.append(
            {
                "来源sheet": line.sheet_name,
                "来源行": line.row_number,
                "门店名称": line.store_name,
                "门店公司代码": line.store_code,
                "开票名称": line.billing_name,
                "二级公司代码": line.match.secondary_code if line.match else "",
                "二级公司名称": line.match.secondary_name if line.match else "",
                "匹配来源": line.match.match_source if line.match else "",
                "分类": line.match.category if line.match else "",
                "发票号码": line.invoice_number,
                "发票文件数": len(invoice_files.get(line.invoice_number, [])),
                "服务费金额": cents_to_amount(line.amount_cents),
                "匹配状态": "已匹配" if line.match else "未匹配",
            }
        )
    write_rows(detail, detail_columns, detail_rows)

    summary_sheet = workbook.create_sheet("二级公司汇总")
    summary_columns = ["二级公司代码", "二级公司名称", "明细行数", "金额合计", "发票数量", "门店数量"]
    write_rows(summary_sheet, summary_columns, aggregate(lines))

    invoices = workbook.create_sheet("发票匹配")
    invoice_columns = ["发票号码", "文件数", "文件名", "关联门店", "二级公司代码", "二级公司名称"]
    invoice_rows = []
    for line in lines:
        files = invoice_files.get(line.invoice_number, [])
        invoice_rows.append(
            {
                "发票号码": line.invoice_number,
                "文件数": len(files),
                "文件名": "; ".join(path.name for path in files),
                "关联门店": line.store_name,
                "二级公司代码": line.match.secondary_code if line.match else "",
                "二级公司名称": line.match.secondary_name if line.match else "",
            }
        )
    write_rows(invoices, invoice_columns, invoice_rows)
    style_workbook(workbook)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 100
    workbook.save(output_path)


def export_import_template(lines: List[FeeLine], output_path: Path) -> None:
    workbook = load_workbook(FINANCE_TEMPLATE)
    sheet = workbook["数据"]
    headers = [cell.value for cell in sheet[1]]
    index = {clean(value): i + 1 for i, value in enumerate(headers) if clean(value)}
    if sheet.max_row > 2:
        sheet.delete_rows(3, sheet.max_row - 2)
    for row_number, row in enumerate([item for item in aggregate(lines) if item["二级公司代码"] != "未匹配"], start=3):
        le_code = row["二级公司代码"]
        values = {
            "leCode": le_code,
            "travelerEmpNo": None,
            "ccCode": f"{le_code}200002",
            "operationSubTypeCode": "hdl0085-001",
            "standardCurrencyAmount": float(row["金额合计"]),
            "taxType": "001105",
            "assetsMainNo": None,
            "assetsChildNo": None,
            "assetsCount": None,
            "vendorCode": None,
        }
        for key, value in values.items():
            column = index.get(key)
            if column:
                sheet.cell(row=row_number, column=column, value=value)
    workbook.save(output_path)


def main() -> None:
    by_code, by_name = load_structure_matches(MATCH_TABLE)
    lines = parse_fee_lines(SOURCE_FEE_FILE, by_code, by_name)
    invoice_files = map_invoice_files(INVOICE_ROOT)
    copied_count = organize_invoices(lines, invoice_files, OUTPUT_INVOICES)
    export_review_workbook(lines, invoice_files, OUTPUT_REVIEW)
    export_import_template(lines, OUTPUT_IMPORT)
    counts = Counter(line.match.secondary_code if line.match else "未匹配" for line in lines)
    print(
        {
            "lines": len(lines),
            "matched": sum(1 for line in lines if line.match),
            "unmatched": sum(1 for line in lines if not line.match),
            "secondary_company_count": len(counts) - (1 if "未匹配" in counts else 0),
            "invoice_files_copied": copied_count,
            "review_file": str(OUTPUT_REVIEW),
            "import_file": str(OUTPUT_IMPORT),
            "invoice_dir": str(OUTPUT_INVOICES),
        }
    )


if __name__ == "__main__":
    main()
