from __future__ import annotations

import tempfile
import unittest
import zipfile
from pathlib import Path

from finance_agent.adapter import MockFinanceSystemAdapter
from finance_agent.config import DEFAULT_CONFIG
from finance_agent.engine import process_batch
from finance_agent.parser import cents_to_amount
from finance_agent.server import _expand_invoice_uploads
from finance_agent.template import export_upload_template_xlsx


def make_workbook(path: Path, rows: list[list[object]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "服务费"
    sheet.append(["二级公司", "季度", "门店代码", "门店名称", "合同编号", "发票号码", "服务费"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def make_haidilao_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    payment = workbook.active
    payment.title = "付款明细1"
    payment.append(["门店", "2026年1-3月服务费", "开票名称", "发票号码"])
    payment.append([None, None, None, None])
    payment.append(["北京二店", 100, "海鸿达（北京）餐饮管理有限公司", "12345678"])
    payment.append(["北京十九店", 200, "海鸿达（北京）餐饮管理有限公司", "87654321"])
    payment.append(["北京五十九店", 300, "海鸿达（北京）餐饮管理有限公司", "11112222"])

    company = workbook.create_sheet("所属公司")
    company.append(["简称", "母公司"])
    company.append(["北京二店", 1044])
    company.append(["北京十九店", 1044])

    stores = workbook.create_sheet("火锅门店")
    stores.append(["店名", "公司代码", "开票名称"])
    stores.append(["北京二店", "BJ02", "海鸿达（北京）餐饮管理有限公司"])
    stores.append(["北京十九店", 1119, "海鸿达（北京）餐饮管理有限公司"])
    stores.append(["北京五十九店", 1159, "海鸿达（北京）餐饮管理有限公司"])
    workbook.save(path)


def make_finance_import_template(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        [
            "leCode",
            "travelerEmpNo",
            "ccCode",
            "operationSubTypeCode",
            "standardCurrencyAmount",
            "taxType",
            "assetsMainNo",
            "assetsChildNo",
            "assetsCount",
            "vendorCode",
        ]
    )
    sheet.append(["示例", None, None, None, None, None, None, None, None, None])
    workbook.save(path)


def make_structure_matching_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "账单主体匹配"
    sheet.append(["分类大类", "公司代码", "门店/主体简称", "二级公司代码", "二级公司名称"])
    sheet.append(["火锅门店", "BJ02", "北京二店", "1003", "海鸿达（北京）餐饮管理有限公司"])
    sheet.append(["火锅门店", "1119", "北京十九店", "1003", "海鸿达（北京）餐饮管理有限公司"])
    sheet.append(["火锅门店", "1159", "北京五十九店", "1003", "海鸿达（北京）餐饮管理有限公司"])
    workbook.save(path)


class FinanceAgentTests(unittest.TestCase):
    def test_process_batch_groups_by_company_quarter(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "service_fee.xlsx"
            make_workbook(
                source,
                [
                    ["华东二级公司", "2026 Q1", "S001", "上海一店", "C-001", "12345678", 1000],
                    ["华东二级公司", "2026 Q1", "S002", "上海二店", "C-002", "87654321", 2500.5],
                    ["华南二级公司", "2026 Q1", "S003", "广州一店", "C-003", "11112222", "300.20"],
                ],
            )
            invoices = [
                root / "发票_12345678.pdf",
                root / "发票_87654321.pdf",
                root / "发票_11112222.pdf",
            ]
            for invoice in invoices:
                invoice.write_text("sample", encoding="utf-8")

            batch = process_batch(
                batch_id="batch-ok",
                name="2026 Q1 服务费",
                source_file=source,
                invoice_files=invoices,
                template_file=None,
                quarter="2026 Q1",
                created_by="tester",
                config=DEFAULT_CONFIG,
            )

            self.assertEqual(batch.status, "ready")
            self.assertEqual(batch.line_count, 3)
            self.assertEqual(batch.company_count, 2)
            east = next(bill for bill in batch.bills if bill.secondary_company == "华东二级公司")
            self.assertEqual(cents_to_amount(east.total_amount_cents), "3500.50")
            self.assertEqual(len([match for match in batch.invoice_matches if match.status == "matched"]), 3)

    def test_missing_invoice_file_blocks_submission(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "service_fee.xlsx"
            make_workbook(
                source,
                [["华东二级公司", "2026 Q1", "S001", "上海一店", "C-001", "12345678", 1000]],
            )

            batch = process_batch(
                batch_id="batch-missing-invoice",
                name="2026 Q1 服务费",
                source_file=source,
                invoice_files=[],
                template_file=None,
                quarter="2026 Q1",
                created_by="tester",
                config=DEFAULT_CONFIG,
            )

            self.assertEqual(batch.status, "blocked")
            self.assertTrue(any(issue.code == "invoice_file_missing" for issue in batch.issues))

    def test_duplicate_invoice_file_blocks_submission(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "service_fee.xlsx"
            make_workbook(
                source,
                [["华东二级公司", "2026 Q1", "S001", "上海一店", "C-001", "12345678", 1000]],
            )
            invoices = [root / "发票_12345678_A.pdf", root / "发票_12345678_B.pdf"]
            for invoice in invoices:
                invoice.write_text("sample", encoding="utf-8")

            batch = process_batch(
                batch_id="batch-dup-invoice",
                name="2026 Q1 服务费",
                source_file=source,
                invoice_files=invoices,
                template_file=None,
                quarter="2026 Q1",
                created_by="tester",
                config=DEFAULT_CONFIG,
            )

            self.assertEqual(batch.status, "blocked")
            self.assertTrue(any(issue.code == "duplicate_invoice_file" for issue in batch.issues))

    def test_adapter_submits_ready_batch_and_export_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "service_fee.xlsx"
            make_workbook(
                source,
                [
                    ["华东二级公司", "2026 Q1", "S001", "上海一店", "C-001", "12345678", 1000],
                    ["华南二级公司", "2026 Q1", "S003", "广州一店", "C-003", "11112222", 300],
                ],
            )
            invoices = [root / "发票_12345678.pdf", root / "发票_11112222.pdf"]
            for invoice in invoices:
                invoice.write_text("sample", encoding="utf-8")

            batch = process_batch(
                batch_id="batch-submit",
                name="2026 Q1 服务费",
                source_file=source,
                invoice_files=invoices,
                template_file=None,
                quarter="2026 Q1",
                created_by="tester",
                config=DEFAULT_CONFIG,
            )
            template_bytes = export_upload_template_xlsx(batch, DEFAULT_CONFIG)
            submissions = MockFinanceSystemAdapter(DEFAULT_CONFIG).submit_batch(batch, "tester")

            self.assertGreater(len(template_bytes), 1000)
            self.assertEqual(batch.status, "submitted")
            self.assertEqual(len(submissions), 2)

    def test_haidilao_parent_company_sheet_drives_aggregation_and_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "haidilao.xlsx"
            template = root / "finance_template.xlsx"
            make_haidilao_workbook(source)
            make_finance_import_template(template)
            invoices = [root / "发票_12345678.pdf", root / "发票_87654321.pdf", root / "发票_11112222.pdf"]
            for invoice in invoices:
                invoice.write_text("sample", encoding="utf-8")
            organized_dir = root / "organized"

            batch = process_batch(
                batch_id="batch-parent",
                name="2026 Q1 服务费",
                source_file=source,
                invoice_files=invoices,
                template_file=template,
                quarter="2026 Q1",
                created_by="tester",
                config=DEFAULT_CONFIG,
                organized_invoice_dir=organized_dir,
            )

            self.assertEqual(batch.status, "ready")
            self.assertEqual(batch.company_count, 1)
            bill = batch.bills[0]
            self.assertEqual(bill.secondary_company, "1044 海鸿达（北京）餐饮管理有限公司")
            self.assertEqual(cents_to_amount(bill.total_amount_cents), "600.00")
            self.assertEqual(bill.store_codes, ["1119", "1159", "BJ02"])
            self.assertEqual(len(list(organized_dir.rglob("*.pdf"))), 3)

            from openpyxl import load_workbook

            output = root / "output.xlsx"
            output.write_bytes(export_upload_template_xlsx(batch, DEFAULT_CONFIG))
            sheet = load_workbook(output, data_only=True)["数据"]
            self.assertEqual(str(sheet["A3"].value), "1044")
            self.assertEqual(str(sheet["C3"].value), "1044200002")
            self.assertEqual(float(sheet["E3"].value), 600.0)

    def test_company_structure_file_remaps_to_secondary_company_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "haidilao.xlsx"
            template = root / "finance_template.xlsx"
            structure = root / "company_structure.xlsx"
            make_haidilao_workbook(source)
            make_finance_import_template(template)
            make_structure_matching_workbook(structure)
            invoices = [root / "发票_12345678.pdf", root / "发票_87654321.pdf", root / "发票_11112222.pdf"]
            for invoice in invoices:
                invoice.write_text("sample", encoding="utf-8")
            organized_dir = root / "organized"
            workbook_dir = root / "company_workbooks"

            batch = process_batch(
                batch_id="batch-secondary",
                name="2026 Q1 服务费",
                source_file=source,
                invoice_files=invoices,
                template_file=template,
                quarter="2026 Q1",
                created_by="tester",
                config=DEFAULT_CONFIG,
                structure_file=structure,
                organized_invoice_dir=organized_dir,
                company_workbook_dir=workbook_dir,
            )

            self.assertEqual(batch.status, "ready")
            self.assertEqual(batch.company_count, 1)
            bill = batch.bills[0]
            self.assertEqual(bill.secondary_company, "1003 海鸿达（北京）餐饮管理有限公司")
            self.assertEqual(cents_to_amount(bill.total_amount_cents), "600.00")
            self.assertEqual({line.secondary_company for line in batch.lines}, {bill.secondary_company})
            self.assertEqual(len(list(organized_dir.rglob("*.pdf"))), 3)
            self.assertEqual(len(list(workbook_dir.glob("*.xlsx"))), 1)

            from openpyxl import load_workbook

            output = root / "output.xlsx"
            output.write_bytes(export_upload_template_xlsx(batch, DEFAULT_CONFIG))
            sheet = load_workbook(output, data_only=True)["数据"]
            self.assertEqual(str(sheet["A3"].value), "1003")
            self.assertEqual(str(sheet["C3"].value), "1003200002")
            self.assertEqual(float(sheet["E3"].value), 600.0)

    def test_invoice_zip_upload_expands_nested_invoice_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            zip_path = root / "invoices.zip"
            with zipfile.ZipFile(zip_path, "w") as archive:
                archive.writestr("华东/一月/服务费发票_12345678.pdf", "sample")
                archive.writestr("华南/服务费发票_11112222.ofd", "sample")
                archive.writestr("说明.txt", "ignore me")
                archive.writestr("__MACOSX/._服务费发票_99999999.pdf", "ignore me")

            expanded = _expand_invoice_uploads([zip_path], root / "expanded")

            self.assertEqual(sorted(path.name for path in expanded), ["服务费发票_11112222.ofd", "服务费发票_12345678.pdf"])


if __name__ == "__main__":
    unittest.main()
